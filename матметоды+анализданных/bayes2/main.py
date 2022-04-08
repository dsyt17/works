from openpyxl import load_workbook

book = load_workbook(filename='data.xlsx')
sheet = book['Лист1']

sky = ['солнечно', 'пасмурно', 'дождь']
t = ['жарко', 'тепло', 'прохладно']
wet = ['высокая', 'нормальная']
wind = ['умененный', 'сильный']
status = ['', 'НЕТ', 'ДА']
counter = 1;

for y in (1, 2):

    p_y = float(sheet.cell(row=11, column=3 - y).value)

    for x1 in range(3):

        p_x1_y = float(sheet.cell(row=1 + x1, column=3 - y).value)
        p_x1 = float(sheet.cell(row=1 + x1, column=3).value)

        for x2 in range(3):

            p_x2_y = float(sheet.cell(row=4 + x2, column=3 - y).value)
            p_x2 = float(sheet.cell(row=4 + x2, column=3).value)

            for x3 in range(2):

                p_x3_y = float(sheet.cell(row=7 + x3, column=3 - y).value)
                p_x3 = float(sheet.cell(row=7 + x3, column=3).value)

                for x4 in range(2):
                    p_x4_y = float(sheet.cell(row=9 + x4, column=3 - y).value)
                    p_x4 = float(sheet.cell(row=9 + x4, column=3).value)

                    result = (p_x1_y * p_x2_y * p_x3_y * p_x4_y * p_y) / (p_x1 * p_x2 * p_x3 * p_x4)

                    if result == 0:
                        result = ((p_x1_y + 1) * (p_x2_y + 1) * (p_x3_y + 1) * (p_x4_y + 1) * p_y) / (
                                (p_x1 + 1) * (p_x2 + 1) * (p_x3 + 1) * (p_x4 + 1))

                    if result > 1:
                        result = 1

                    print(str(counter) + '.', status[y], '|', sky[x1], t[x2], wet[x3], wind[x4], '|', ("%.5f" % result))
                    counter += 1

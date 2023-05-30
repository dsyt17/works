from openpyxl import load_workbook

book = load_workbook(filename='irisy.xlsx')
l = book['l']
r = book['r']

answer = "irisyAnswer.xlsx"

for i in range(2, 32):
    for j in range(2, 122):
        l.cell(row=j, column=7).value = abs(float(l.cell(j, 3).value) - float(r.cell(i, 3).value)) + abs(
            float(l.cell(j, 4).value) - float(r.cell(i, 4).value))

    _min = l.cell(2, 7).value
    n_min = 2

    for j in range(3, 122):
        if l.cell(j, 7).value < _min:
            _min = l.cell(j, 7).value
            n_min = j

    l.cell(1, 7).value = _min
    l.cell(1, 9).value = n_min

    r.cell(i, 7).value = l.cell(n_min, 5).value

book.save(answer)
print("Книга сохранена как " + answer)
